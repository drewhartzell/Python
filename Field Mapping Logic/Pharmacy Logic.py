#!/usr/bin/env python3
import sys
import pandas as pd
from openpyxl import load_workbook
import os

# Check for command line arguments (expecting 4 arguments: template, data, vendor code)
if len(sys.argv) != 4:
    sys.exit("Usage: python mapping_script.py <template_file> <data_file> <vendor_file_code>")

# Get file paths and vendor file code from command-line arguments
template_file = sys.argv[1]
data_file = sys.argv[2]
vendor_code = sys.argv[3]

# Determine file type and load the healthcare data
file_extension = data_file.split('.')[-1].lower()

if file_extension == "csv":
    file = pd.read_csv(data_file)
# elif file_extension in ["xls", "xlsx"]:
    # file = pd.read_excel(data_file)
elif file_extension in ["xlsx", "xls"]:
    try:
        # Load the first sheet of the Excel file
        file = pd.read_excel(data_file, sheet_name=0, dtype=str)

        # Remove leading/trailing spaces from column names
        file.columns = file.columns.str.strip()
        
        # If empty, raise an error
        if file.empty:
            sys.exit("Error: The selected Excel file is empty.")

    except Exception as e:
        sys.exit(f"Error reading Excel file: {e}")
   
# elif file_extension == "txt":
    # file = pd.read_csv(data_file, delimiter="\t")  # Assumes tab-delimited; adjust if needed
elif file_extension == "txt":
    try:
        # Attempt to automatically detect delimiter
        with open(data_file, 'r', encoding='utf-8') as f:
            first_line = f.readline()
            # Check for common delimiters
            if '\t' in first_line:
                delimiter = '\t'  # Tab-separated
            elif ',' in first_line:
                delimiter = ','  # Comma-separated
            elif '|' in first_line:
                delimiter = '|'  # Pipe-separated
            elif ' ' in first_line:
                delimiter = ' '  # Space-separated (not ideal)
            else:
                delimiter = '\t'  # Default to tab

        # Read file with detected delimiter
        file = pd.read_csv(data_file, delimiter=delimiter)
    except Exception as e:
        sys.exit(f"Error reading TXT file: {e}")

else:
    sys.exit(f"Unsupported file format: {file_extension}")

# Initialize an empty dictionary to store mappings
data_mapping = {}

# Define a function to map values to columns
def map_values_to_columns(column_name, values, mapping):
    mapping[column_name] = 'Error'  # Default to 'Error' if no mapping is found

    if column_name not in mapping:
        mapping[column_name] = column_name  # Default to the same column name


        ### Subscriber Info



    if (any(substring in column_name.lower() for substring in ['first', 'firstname','first name','fname'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(".", "").replace("'", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber First Name'
# Encompasses values that have middle initials and apostrophies within the first name when applicable

    elif (any(substring in column_name.lower() for substring in ['last', 'lastname','last name','lname'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace("'", "").replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber Last Name'
# Encompasses values that have hyphenated last names and apostrophies within the last name when applicable
                
    elif (any(substring in column_name.lower() for substring in ['middle', 'initial','mid','init','mname'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(".", "").replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber Middle Name'
# Encompasses values that have hyphenated middle names and initials within the middle name when applicable
                
    elif (any(substring in column_name.lower() for substring in ['prefix'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(".", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber Name Prefix'
# Encompasses values that have periods within the prefix
                
    elif (any(substring in column_name.lower() for substring in ['suffix'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(".", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Subscriber Name Suffix'
# Encompasses values that have periods, alpha characters or integers within the prefix
                
    elif (any(substring in column_name.lower() for substring in ['ssn','social'])
        and (any(substring in column_name.lower() for substring in ['sub','emp','subscriber','employee','EE']))
            and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Subscriber SSN'
# Encompasses values that have only integers or hyphenated SSNs
                
    elif (any(substring in column_name.lower() for substring in ['DOB','birth','bday'])
        and (any(substring in column_name.lower() for substring in ['sub','emp','subscriber','employee','EE']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Subscriber Date of Birth'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes
                
    elif (any(substring in column_name.lower() for substring in ['gender','sex'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Subscriber Gender'
# Encompasses values of 'F', 'Female', or '02' alpha characters or integers within gender codes

        ## Demo Info

    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and (any(substring in column_name.lower() for substring in ['2']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Subscriber Address 2'        
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Subscriber Address'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs

    elif (any(substring in column_name.lower() for substring in ['city'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber City'
# Encompasses values that are only alpha characters and hyphenated cities
                
    elif (any(substring in column_name.lower() for substring in ['state','st'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Subscriber State'
# Encompasses values that are only alpha characters
                
    elif (any(substring in column_name.lower() for substring in ['zip','zcode'])
        and (any(substring in column_name.lower() for substring in ['sub','emp','subscriber','employee','EE']))
            and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Subscriber Zip Code'
# Encompasses values that are 5 digits long or hyphenated 9 digit zip codes



        ### Member Info
                
                

    elif (any(substring in column_name.lower() for substring in ['first', 'firstname','first name','fname'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(".", "").replace("'", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient First Name'
# Encompasses values that have middle initials within the first name when applicable and apostrophies with the first name
                
    elif (any(substring in column_name.lower() for substring in ['last', 'lastname','last name','lname'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace("'", "").replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient Last Name'
# Encompasses values that have hyphenated last names and apostrophies within the last name when applicable
                
    elif (any(substring in column_name.lower() for substring in ['middle', 'initial','mid','init','mname'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(".", "").replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient Middle Name'
# Encompasses values that have hyphenated middle names and initials within the middle name when applicable
                
    elif (any(substring in column_name.lower() for substring in ['prefix'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(".", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient Name Prefix'
# Encompasses values that have periods within the prefix
                
    elif (any(substring in column_name.lower() for substring in ['suffix'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(".", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Patient Name Suffix'
# Encompasses values that have periods, alpha characters or integers within the prefix
                
    elif (any(substring in column_name.lower() for substring in ['ssn','social'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Patient SSN'
# Encompasses values that have only integers or hyphenated SSNs
                
    elif (any(substring in column_name.lower() for substring in ['id'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace("-", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Patient ID (Vendor)'
# Encompasses values that are alphanumberical and have hyphens when applicable
                
    elif (any(substring in column_name.lower() for substring in ['DOB','birth','bday'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Patient Date of Birth'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes
                
    elif (any(substring in column_name.lower() for substring in ['gender','sex'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Patient Gender'
# Encompasses values of 'F', 'Female', or '02' alpha characters or integers within gender codes
                
        ## Demo Info

    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and (any(substring in column_name.lower() for substring in ['2']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Patient Address 2' 
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Patient Address'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
        
    elif (any(substring in column_name.lower() for substring in ['city'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace("-", "").replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient City'
# Encompasses values that are only alpha characters and hyphenated cities
                
    elif (any(substring in column_name.lower() for substring in ['state','st'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace(" ", "").isalpha()) for value in values)):
                mapping[column_name] = 'Patient State'
# Encompasses values that are only alpha characters

    elif (any(substring in column_name.lower() for substring in ['zip','zcode'])
        and (any(substring in column_name.lower() for substring in ['mem', 'member','dep','dependent','patient','pat']))
            and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Patient Zip Code'
# Encompasses values that are 5 digits long or hyphenated 9 digit zip codes



## Billing Provider Info
                


    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and (any(substring in column_name.lower() for substring in ['2']))
                    and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                        mapping[column_name] = 'Prescriber Address 2' 
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Prescriber Address'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
        
    elif (any(substring in column_name.lower() for substring in ['city'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace("-", "").replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Prescriber City'
# Encompasses values that are only alpha characters and hyphenated cities
                
    elif (any(substring in column_name.lower() for substring in ['state','st'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Prescriber State'
# Encompasses values that are only alpha characters
                    
    elif (any(substring in column_name.lower() for substring in ['zip','zcode'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                    mapping[column_name] = 'Prescriber Zip Code'
# Encompasses values that are 5 digits long or hyphenated 9 digit zip codes
                    
    elif (any(substring in column_name.lower() for substring in ['npi','npidentifier'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace("-", "").replace(" ", "").isdigit() for value in values))):
                    mapping[column_name] = 'Prescriber NPI'
# Encompasses values that are alphanumberical and have hyphens when applicable

    elif (any(substring in column_name.lower() for substring in ['name','nm'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace("-", "").replace(",", "").replace(" ", "").isalpha() for value in values))):
                    mapping[column_name] = 'Prescriber Name'
# Encompasses values that are only alpha characters and hyphenated cities   
    
    elif (any(substring in column_name.lower() for substring in ['type'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalpha() for value in values))):
                    mapping[column_name] = 'Prescriber ID Type'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['code','cd','spec','specialty'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Prescriber Specialty'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs    

    elif (any(substring in column_name.lower() for substring in ['id'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pres','prescriber']))
                and all((str(value).replace("-", "").replace(" ", "").isalnum()) or (str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                    mapping[column_name] = 'Prescriber ID'
# Encompasses values that are alphanumberical and have hyphens when applicable                



## Performing Provider Info
                


    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and (any(substring in column_name.lower() for substring in ['2']))
                    and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                        mapping[column_name] = 'Pharmacy Address 2' 
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['addr','address'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Pharmacy Address'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
        
    elif (any(substring in column_name.lower() for substring in ['city'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace("-", "").replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Pharmacy City'
# Encompasses values that are only alpha characters and hyphenated cities
                
    elif (any(substring in column_name.lower() for substring in ['state','st'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Pharmacy State'
# Encompasses values that are only alpha characters
                    
    elif (any(substring in column_name.lower() for substring in ['zip','zcode'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                    mapping[column_name] = 'Pharmacy Zip Code'
# Encompasses values that are 5 digits long or hyphenated 9 digit zip codes
                    
    elif (any(substring in column_name.lower() for substring in ['npi'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace("-", "").replace(" ", "").isdigit()) for value in values)):
                    mapping[column_name] = 'Pharmacy NPI'
# Encompasses values that are alphanumberical and have hyphens when applicable

    elif (any(substring in column_name.lower() for substring in ['name','nm'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace("-", "").replace(",", "").replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Pharmacy Name'
# Encompasses values that are only alpha characters and hyphenated cities   
    
    elif (any(substring in column_name.lower() for substring in ['type'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalpha()) for value in values)):
                    mapping[column_name] = 'Pharmacy ID Type'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['code','cd','spec','specialty'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum()) for value in values)):
                    mapping[column_name] = 'Pharmacy Specialty Code'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs  

    elif (any(substring in column_name.lower() for substring in ['id'])
        and (any(substring in column_name.lower() for substring in ['prov', 'provider']))
            and (any(substring in column_name.lower() for substring in ['pharm','pharmacy','serv','servicing']))
                and all((str(value).replace("-", "").replace(" ", "").isalnum()) for value in values)):
                    mapping[column_name] = 'Pharmacy ID'
# Encompasses values that are alphanumberical and have hyphens when applicable  



        ### Date Info
                    


    elif (any(substring in column_name.lower() for substring in ['service','serv','incurred','inc'])
        and (any(substring in column_name.lower() for substring in ['dt','date']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Date of Service (Incurred)'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes

    elif (any(substring in column_name.lower() for substring in ['paid','adjudicated','pd','adj'])
        and (any(substring in column_name.lower() for substring in ['dt','date']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Date Claim Paid'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes

    elif (any(substring in column_name.lower() for substring in ['discharge','dis'])
        and (any(substring in column_name.lower() for substring in ['dt','date']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Discharge Date'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes   

    elif (any(substring in column_name.lower() for substring in ['prescription','written'])
        and (any(substring in column_name.lower() for substring in ['dt','date']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Date Prescription Written'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes   

    elif (any(substring in column_name.lower() for substring in ['billed','bill'])
        and (any(substring in column_name.lower() for substring in ['dt','date']))
            and all(is_date(value) or (str(value).replace("-", "").replace("/", "").replace(" ", "").isdigit()) for value in values)):
                mapping[column_name] = 'Date Plan Billed'
# Encompasses values that are date fields, string values of only digits, or string values separated by hyphens or slashes   


       ### Procedure/Claim Data
       
       
       
    elif (any(substring in column_name.lower() for substring in ['claim'])
        and (any(substring in column_name.lower() for substring in ['number']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Claim Number'
# Encompasses values that are alphanumberical and have hyphens when applicable
    
    elif (any(substring in column_name.lower() for substring in ['line'])
        and (any(substring in column_name.lower() for substring in ['number']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Line Number'
# Encompasses values that are alphanumberical and have hyphens when applicable



### Pharmacy Claim Data
       
       
       
    elif (any(substring in column_name.lower() for substring in ['formulary'])
        and (any(substring in column_name.lower() for substring in ['status']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Formulary Status'
# Encompasses values that are alphanumberical and have hyphens when applicable
    
    elif (any(substring in column_name.lower() for substring in ['compounded'])
        and (any(substring in column_name.lower() for substring in ['ind']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Compounded Drug Indicator'
# Encompasses values that are alphanumberical and have hyphens when applicable
            
    elif (any(substring in column_name.lower() for substring in ['maintenance','main'])
        and (any(substring in column_name.lower() for substring in ['indicator', 'ind']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Maintenance Drug Indicator'
# Encompasses values that are alphanumberical and have hyphens when applicable
            
    elif (any(substring in column_name.lower() for substring in ['generic'])
        and (any(substring in column_name.lower() for substring in ['indicator', 'ind']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Generic Indicator'
# Encompasses values that are alphanumberical and have hyphens when applicable
                    
    elif (any(substring in column_name.lower() for substring in ['specialty','spec'])
        and (any(substring in column_name.lower() for substring in ['indicator', 'ind']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Specialty Drug Indicator'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs  
                    
    elif (any(substring in column_name.lower() for substring in ['drug','unit'])
        and (any(substring in column_name.lower() for substring in ['measure', 'mod']))
            and (any(substring in column_name.lower() for substring in ['code', 'cd']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Drug Unit Measure Code'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs  

    elif (any(substring in column_name.lower() for substring in ['national'])
        or (any(substring in column_name.lower() for substring in ['ndc']))
        and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
            mapping[column_name] = 'National Drug Code'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs

    elif (any(substring in column_name.lower() for substring in ['rx','pharmacy'])
        and (any(substring in column_name.lower() for substring in ['number','numb']))
            and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Rx Number'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs

    elif (any(substring in column_name.lower() for substring in ['prod','product'])
            and (any(substring in column_name.lower() for substring in ['name', 'nm']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Product Name'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['prod','product'])
            and (any(substring in column_name.lower() for substring in ['strength', 'str']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Product Strength'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['quantity','quan'])
            and (any(substring in column_name.lower() for substring in ['dispensed', 'disp']))
        and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
            mapping[column_name] = 'Quantity Dispensed'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['claim','payment','paid'])
            and (any(substring in column_name.lower() for substring in ['status']))
                and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isalnum() for value in values))):
                    mapping[column_name] = 'Claim Payment Status'
# Encompasses values that are numerical and have periods

    elif (any(substring in column_name.lower() for substring in ['days','day'])
            and (any(substring in column_name.lower() for substring in ['supply', 'sup']))
        and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
            mapping[column_name] = 'Days Supply'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs
                    
    elif (any(substring in column_name.lower() for substring in ['dispense'])
        and (any(substring in column_name.lower() for substring in ['written', 'sup']))
            or (any(substring in column_name.lower() for substring in ['daw'])) 
                and (any(substring in column_name.lower() for substring in ['code','cd']))
        and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
            mapping[column_name] = 'Dispense as Written (DAW)/Product Selection Code'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs

    elif (any(substring in column_name.lower() for substring in ['auth','authorization'])
        and (any(substring in column_name.lower() for substring in ['ind', 'indicator']))
        and all((str(value).replace(".", "").replace("#", "").replace(",", "").replace(" ", "").isdigit() for value in values))):
            mapping[column_name] = 'Prior Authorization Indicator'            
# Encompasses values that are alphanumberical, include periods, commas, and pound signs

    elif (any(substring in column_name.lower() for substring in ['dispenser','dispense'])
        and (any(substring in column_name.lower() for substring in ['type','ty','status']))
            and all((str(value).replace(".", "").replace("#", "").replace("-", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Pharmacy Dispenser Type'
# Encompasses values that are numerical and have periods



        ### Financial Data
                


    elif (any(substring in column_name.lower() for substring in ['coinsurance'])
                and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                    mapping[column_name] = 'Total Coinsurance'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['plan paid'])
        and (any(substring in column_name.lower() for substring in ['amount']))
                and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                    mapping[column_name] = 'Total Amount Plan Paid'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['allowed'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Total Amount Paid by all Sources (Allowed)'
# Encompasses values that are numerical and have periods
    
    elif (any(substring in column_name.lower() for substring in ['deductible'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Total Deductible'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['copay'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Total Copay'
# Encompasses values that are numerical and have periods
                
    elif (any(substring in column_name.lower() for substring in ['ingredient','ingred'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Ingredient Cost Paid'
# Encompasses values that are numerical and have periods
                
    elif (any(substring in column_name.lower() for substring in ['sales'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Sales Tax'
# Encompasses values that are numerical and have periods
                
    elif (any(substring in column_name.lower() for substring in ['dispensing'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Dispensing Fee Paid'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['coordination','COB'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Coordination of Benefits Paid'
# Encompasses values that are numerical and have periods
                    
    elif (any(substring in column_name.lower() for substring in ['submitted'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Total Submitted Amount'
# Encompasses values that are numerical and have periods
    
    elif (any(substring in column_name.lower() for substring in ['unit','wholesale'])
        and (any(substring in column_name.lower() for substring in ['price','amt','amount','paid','pd']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Average Wholesale Unit Price'
# Encompasses values that are numerical and have periods
    
    elif (any(substring in column_name.lower() for substring in ['mac'])
        and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd','price']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'MAC Price'
# Encompasses values that are numerical and have periods

    elif (any(substring in column_name.lower() for substring in ['client'])
        and (any(substring in column_name.lower() for substring in ['basis']))
            and (any(substring in column_name.lower() for substring in ['amount','amt','paid','pd','pricing','price']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Client Pricing Basis of Cost'
# Encompasses values that are numerical and have periods

    elif (any(substring in column_name.lower() for substring in ['vaccine'])
            and (any(substring in column_name.lower() for substring in ['fee','amount','amt','paid','pd','pricing','price']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Vaccine Admin Fee'
# Encompasses values that are numerical and have periods

    elif (any(substring in column_name.lower() for substring in ['u&c'])
            and (any(substring in column_name.lower() for substring in ['fee','amount','amt','paid','pd','pricing','price']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'U&C Amount'
# Encompasses values that are numerical and have periods

    elif (any(substring in column_name.lower() for substring in ['patient','pat'])
            and (any(substring in column_name.lower() for substring in ['daw','penalty']))
                and (any(substring in column_name.lower() for substring in ['fee','amount','amt','paid','pd','pricing','price','cost']))
            and all((str(value).replace(".", "").replace(" ", "").isdigit() for value in values))):
                mapping[column_name] = 'Patient DAW Penalty Cost'
# Encompasses values that are numerical and have periods
        

    
        ### Coverage/Relationship
                


    elif (any(substring in column_name.lower() for substring in ['rel','relationship'])
        and (any(substring in column_name.lower() for substring in ['code', 'tier','cd']))
            and all((str(value).replace("+", "").replace("(", "").replace(")", "").replace("-", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Patient-Subscriber Relationship'
# Encompasses values that are alphanumberical, include paratheses, plus signs, and hyphens

    elif (any(substring in column_name.lower() for substring in ['id'])
        and (any(substring in column_name.lower() for substring in ['sub', 'emp','subscriber','employee','EE']))
            and all((str(value).replace("-", "").replace(" ", "").isalnum() for value in values))):
                mapping[column_name] = 'Subscriber ID (Vendor)'
# Encompasses values that are alphanumberical and have hyphens when applicable

        # If no specific mapping is found, you can leave it as is
    if column_name not in mapping:
        mapping[column_name] = column_name  # Default to the same column name

def is_date(value):
    try:
        pd.to_datetime(value)
        return True
    except (TypeError, ValueError):
        return False

# Iterate over columns and values to map them
for column_name in file.columns:
    # values = file[column_name].unique
    values = file[column_name]
    map_values_to_columns(column_name, values, data_mapping)

# Load the original workbook based on the selected template file
wb = load_workbook(template_file)

# Get the existing 'Field Mapping' sheet
ws = wb['Field Mapping']

# Convert column names to column numbers (A=1, B=2, ...)
column_numbers = {col: idx + 1 for idx, col in enumerate(data_mapping.values())}

# Iterate over the rows in the 'Field Mapping' sheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
    innovu_data_element = row[0].value
    # Check if the innovu data element matches any of the mapped column names
    for column_name in file.columns:
        if innovu_data_element == data_mapping.get(column_name, None):
            column_number = file.columns.get_loc(column_name) + 1  # Adjust to 1-based indexing
            # Write the original header name to column 6 and column number to column 7
            ws.cell(row=row[0].row, column=6).value = column_name
            ws.cell(row=row[0].row, column=7).value = column_number
            break  # Exit the loop once a match is found

# Save the updated workbook with the new mapping, incorporating the vendor file code
#output_filename = f"IA_fm_{vendor_code}__mr_enrD_4.2.xlsx"

# Specify the file path
#file_path = f"C:/Users/andre/Downloads/{output_filename}"

# Save the workbook to the specified location
#wb.save(file_path)

#wb.save(output_filename)


# Generate the output filename using the vendor_code variable
output_filename = f"IA_fm_{vendor_code}_rx_clmD_4.2.xlsx"

# Get the current user's home directory and then the Downloads folder
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")

# Combine the Downloads folder with the output filename to create the full file path
file_path = os.path.join(downloads_folder, output_filename)

# Save the workbook to the specified location
wb.save(file_path)
